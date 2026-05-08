#!/usr/bin/env python3
"""
Generate spoof-resistant AWS Silicon Valley Workflow certificates.

The script keeps the Canva PDF template as the certificate background and
overlays attendee details, a signed verification QR code, and the Rayyan
Shaheer signature block. Each certificate is exported as both PDF and PNG.
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import io
import json
import os
import re
import uuid
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Sequence

import qrcode
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey, Ed25519PublicKey
from pdf2image import convert_from_bytes
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime with a clearer message.
    load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = PROJECT_ROOT / "assets" / "templates" / "aws-session-attendee-certificate-template.pdf"
DEFAULT_LOGO = PROJECT_ROOT / "assets" / "logos" / "aws-cc-fast-pwr-black-fg-no-bg.png"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs" / "certificates"
DEFAULT_KEYS_DIR = PROJECT_ROOT / ".certificate_keys"
DEFAULT_PRIVATE_KEY = DEFAULT_KEYS_DIR / "private_key.pem"
DEFAULT_PUBLIC_KEY = DEFAULT_KEYS_DIR / "public_key.pem"
DEFAULT_EVENT_TITLE = "AWS Silicon Valley Workflow"
DEFAULT_SESSIONS_TOTAL = "5"
DEFAULT_SIGNER_NAME = "Rayyan Shaheer"
DEFAULT_SIGNER_TITLE = "AWS Cloud Club Captain at FAST Peshawar"
DEFAULT_CERTIFICATE_TYPE = "attendance"

BROWN = colors.HexColor("#a36532")
LIGHT_BROWN = colors.HexColor("#c78555")
BG_FILL = colors.HexColor("#fbfaf6")
INK = colors.HexColor("#1f2933")

NAME_COLUMNS = (
    "name",
    "full name",
    "full_name",
    "attendee",
    "attendee name",
    "attendee_name",
    "participant",
    "participant name",
    "participant_name",
)
ROLLNO_COLUMNS = (
    "rollno",
    "roll no",
    "roll_no",
    "roll number",
    "roll_number",
    "registration no",
    "registration_no",
    "student id",
    "student_id",
)
EMAIL_COLUMNS = ("email", "email address", "email_address")
EVENT_COLUMNS = ("event", "event title", "event_title", "session", "session title", "session_title")


@dataclass(frozen=True)
class Attendee:
    name: str
    rollno: str
    email: str
    event_title: str | None = None


@dataclass(frozen=True)
class SignedCertificate:
    attendee: Attendee
    cert_id: str
    key_id: str
    token: str
    verify_url: str
    payload: dict[str, str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate signed AWS Silicon Valley Workflow certificates from CSV/XLSX."
    )
    parser.add_argument(
        "--input",
        "--csv",
        dest="input_path",
        required=True,
        type=Path,
        help="CSV or XLSX file with name, rollno, and email columns.",
    )
    parser.add_argument(
        "--verify-base-url",
        required=True,
        help="Verification page URL, for example https://rayyanshaheer.com/verify.",
    )
    parser.add_argument(
        "--template",
        default=DEFAULT_TEMPLATE,
        type=Path,
        help="Certificate PDF template.",
    )
    parser.add_argument(
        "--logo",
        default=DEFAULT_LOGO,
        type=Path,
        help="Replacement logo image.",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_DIR,
        type=Path,
        help="Folder for generated certificates.",
    )
    parser.add_argument(
        "--event",
        default=DEFAULT_EVENT_TITLE,
        help="Default event title used when input rows do not override it.",
    )
    parser.add_argument(
        "--sessions-total",
        "--workshops",
        dest="sessions_total",
        default=DEFAULT_SESSIONS_TOTAL,
        help="Total event sessions completed. --workshops is kept as a legacy alias.",
    )
    parser.add_argument(
        "--signer-name",
        default=DEFAULT_SIGNER_NAME,
        help="Name shown in the bottom-right signature block.",
    )
    parser.add_argument(
        "--signer-title",
        default=DEFAULT_SIGNER_TITLE,
        help="Title shown in the bottom-right signature block.",
    )
    parser.add_argument(
        "--certificate-type",
        default=DEFAULT_CERTIFICATE_TYPE,
        help="Certificate type included in the signed payload.",
    )
    parser.add_argument(
        "--issue-date",
        default=date.today().isoformat(),
        help="Issue date included in the signed payload, formatted YYYY-MM-DD.",
    )
    parser.add_argument(
        "--private-key",
        default=DEFAULT_PRIVATE_KEY,
        type=Path,
        help="Local Ed25519 private signing key. Created automatically if missing.",
    )
    parser.add_argument(
        "--public-key",
        default=DEFAULT_PUBLIC_KEY,
        type=Path,
        help="Public key file exported for the verification website.",
    )
    parser.add_argument(
        "--combined-name",
        default="all_certificates.pdf",
        help="File name for the combined multi-page PDF.",
    )
    parser.add_argument(
        "--uppercase-names",
        action="store_true",
        help="Render attendee names in uppercase.",
    )
    return parser.parse_args()


def normalized_key(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").strip().lower()).strip()


def normalize_base_url(value: str) -> str:
    cleaned = value.strip()
    if not cleaned:
        raise ValueError("--verify-base-url is required.")
    return cleaned.rstrip("/")


def find_column(fieldnames: Sequence[str], candidates: Sequence[str]) -> str | None:
    field_lookup = {normalized_key(field): field for field in fieldnames}
    return next((field_lookup.get(normalized_key(candidate)) for candidate in candidates if normalized_key(candidate) in field_lookup), None)


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_to_attendee(
    row: dict[str, object],
    name_field: str,
    rollno_field: str,
    email_field: str,
    event_field: str | None,
    line_number: int,
) -> Attendee | None:
    name = cell_to_text(row.get(name_field))
    rollno = cell_to_text(row.get(rollno_field))
    email = cell_to_text(row.get(email_field))
    if not any((name, rollno, email)):
        return None
    missing = [
        label
        for label, value in (("name", name), ("rollno", rollno), ("email", email))
        if not value
    ]
    if missing:
        raise ValueError(f"Row {line_number} is missing required value(s): {', '.join(missing)}")
    event_title = cell_to_text(row.get(event_field)) if event_field else ""
    return Attendee(name=name, rollno=rollno, email=email, event_title=event_title or None)


def require_columns(fieldnames: Sequence[str]) -> tuple[str, str, str, str | None]:
    name_field = find_column(fieldnames, NAME_COLUMNS)
    rollno_field = find_column(fieldnames, ROLLNO_COLUMNS)
    email_field = find_column(fieldnames, EMAIL_COLUMNS)
    event_field = find_column(fieldnames, EVENT_COLUMNS)
    missing = [
        label
        for label, field in (("name", name_field), ("rollno", rollno_field), ("email", email_field))
        if not field
    ]
    if missing:
        available = ", ".join(fieldnames)
        raise ValueError(
            f"Input is missing required column(s): {', '.join(missing)}. Available columns: {available}"
        )
    return name_field, rollno_field, email_field, event_field


def read_csv_attendees(input_path: Path) -> list[Attendee]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV must include a header row with name, rollno, and email columns.")
        name_field, rollno_field, email_field, event_field = require_columns(reader.fieldnames)

        attendees: list[Attendee] = []
        for line_number, row in enumerate(reader, start=2):
            attendee = row_to_attendee(row, name_field, rollno_field, email_field, event_field, line_number)
            if attendee:
                attendees.append(attendee)
        return attendees


def read_xlsx_attendees(input_path: Path) -> list[Attendee]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for .xlsx input. Install dependencies from requirements.txt.")

    workbook = load_workbook(input_path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration as exc:
        raise ValueError("XLSX file is empty.") from exc

    fieldnames = [cell_to_text(value) for value in header_values]
    if not any(fieldnames):
        raise ValueError("XLSX must include a header row with name, rollno, and email columns.")
    name_field, rollno_field, email_field, event_field = require_columns(fieldnames)

    attendees: list[Attendee] = []
    for line_number, values in enumerate(rows, start=2):
        row = {field: values[index] if index < len(values) else "" for index, field in enumerate(fieldnames)}
        attendee = row_to_attendee(row, name_field, rollno_field, email_field, event_field, line_number)
        if attendee:
            attendees.append(attendee)
    return attendees


def read_attendees(input_path: Path) -> list[Attendee]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        attendees = read_csv_attendees(input_path)
    elif suffix == ".xlsx":
        attendees = read_xlsx_attendees(input_path)
    elif suffix == ".xls":
        raise ValueError("Legacy .xls files are not supported. Save the file as .xlsx or .csv.")
    else:
        raise ValueError("Input must be a .csv or .xlsx file.")

    if not attendees:
        raise ValueError("No attendee rows were found in the input file.")
    return attendees


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._")
    return cleaned or "certificate"


def certificate_filename(attendee: Attendee) -> str:
    return f"{safe_filename(attendee.rollno)}_{safe_filename(attendee.name)}"


def b64url_encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def load_or_create_private_key(private_key_path: Path, public_key_path: Path) -> Ed25519PrivateKey:
    private_key_path.parent.mkdir(parents=True, exist_ok=True)
    if private_key_path.exists():
        private_key = serialization.load_pem_private_key(private_key_path.read_bytes(), password=None)
        if not isinstance(private_key, Ed25519PrivateKey):
            raise TypeError(f"Private key is not an Ed25519 key: {private_key_path}")
    else:
        private_key = Ed25519PrivateKey.generate()
        private_key_path.write_bytes(
            private_key.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.PKCS8,
                encryption_algorithm=serialization.NoEncryption(),
            )
        )
        os.chmod(private_key_path, 0o600)

    public_key_path.parent.mkdir(parents=True, exist_ok=True)
    public_key_path.write_bytes(
        private_key.public_key().public_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PublicFormat.SubjectPublicKeyInfo,
        )
    )
    return private_key


def key_id_for_public_key(public_key: Ed25519PublicKey) -> str:
    public_der = public_key.public_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PublicFormat.SubjectPublicKeyInfo,
    )
    return hashlib.sha256(public_der).hexdigest()[:16]


def public_key_pem(public_key: Ed25519PublicKey) -> str:
    return public_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo,
    ).decode("ascii")


def email_hash(email: str) -> str:
    return hashlib.sha256(email.strip().lower().encode("utf-8")).hexdigest()


def make_cert_id(issue_date: str) -> str:
    return f"AWS-SVW-{issue_date.replace('-', '')}-{uuid.uuid4().hex[:10].upper()}"


def signed_token(payload: dict[str, str], private_key: Ed25519PrivateKey) -> str:
    payload_json = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    payload_part = b64url_encode(payload_json)
    signature = private_key.sign(payload_part.encode("ascii"))
    signature_part = b64url_encode(signature)
    return f"v1.{payload_part}.{signature_part}"


def sign_certificate(
    attendee: Attendee,
    private_key: Ed25519PrivateKey,
    key_id: str,
    verify_base_url: str,
    default_event_title: str,
    sessions_total: str,
    signer_name: str,
    signer_title: str,
    certificate_type: str,
    issue_date: str,
) -> SignedCertificate:
    cert_id = make_cert_id(issue_date)
    payload = {
        "cert_id": cert_id,
        "recipient_name": attendee.name,
        "rollno": attendee.rollno,
        "email_sha256": email_hash(attendee.email),
        "event": attendee.event_title or default_event_title,
        "sessions_total": str(sessions_total),
        "issuer_name": signer_name,
        "issuer_title": signer_title,
        "issue_date": issue_date,
        "certificate_type": certificate_type,
        "key_id": key_id,
    }
    token = signed_token(payload, private_key)
    verify_url = f"{verify_base_url}#token={token}"
    return SignedCertificate(
        attendee=attendee,
        cert_id=cert_id,
        key_id=key_id,
        token=token,
        verify_url=verify_url,
        payload=payload,
    )


def crop_transparent_border(image: Image.Image) -> Image.Image:
    image = image.convert("RGBA")
    alpha = image.getchannel("A")
    bbox = alpha.getbbox()
    return image.crop(bbox) if bbox else image


def find_signature_font() -> Path | None:
    candidates = (
        Path("/usr/share/fonts/opentype/urw-base35/Z003-MediumItalic.otf"),
        Path("/usr/share/fonts/opentype/linux-libertine/LinLibertine_RI.otf"),
        Path("/usr/share/fonts/opentype/urw-base35/NimbusRoman-Italic.otf"),
    )
    return next((path for path in candidates if path.exists()), None)


def make_signature_image(name: str) -> Image.Image:
    font_path = find_signature_font()
    width, height = 1000, 260
    image = Image.new("RGBA", (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(image)

    if font_path:
        font = ImageFont.truetype(str(font_path), 120)
    else:
        font = ImageFont.load_default()

    bbox = draw.textbbox((0, 0), name, font=font)
    text_w = bbox[2] - bbox[0]
    text_h = bbox[3] - bbox[1]
    x = max(20, (width - text_w) // 2)
    y = max(8, (height - text_h) // 2 - 8)
    draw.text((x, y), name, font=font, fill=(24, 24, 24, 235))

    cropped = crop_transparent_border(image)
    padding = 18
    padded = Image.new(
        "RGBA",
        (cropped.width + padding * 2, cropped.height + padding * 2),
        (255, 255, 255, 0),
    )
    padded.alpha_composite(cropped, (padding, padding))
    return padded


def make_qr_image(url: str) -> Image.Image:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
    )
    qr.add_data(url)
    qr.make(fit=True)
    return qr.make_image(fill_color="#1f2933", back_color="#fbfaf6").convert("RGB")


def fit_font_size(
    text: str,
    font_name: str,
    max_width: float,
    start_size: int,
    min_size: int,
) -> int:
    for size in range(start_size, min_size - 1, -1):
        if stringWidth(text, font_name, size) <= max_width:
            return size
    return min_size


def draw_centered_fit(
    pdf: canvas.Canvas,
    text: str,
    x_center: float,
    y: float,
    max_width: float,
    font_name: str,
    start_size: int,
    min_size: int,
    color: colors.Color,
) -> None:
    size = fit_font_size(text, font_name, max_width, start_size, min_size)
    pdf.setFillColor(color)
    pdf.setFont(font_name, size)
    pdf.drawCentredString(x_center, y, text)


def draw_wrapped_centered(
    pdf: canvas.Canvas,
    text: str,
    x_center: float,
    first_y: float,
    max_width: float,
    font_name: str,
    font_size: int,
    color: colors.Color,
    line_gap: float = 12,
) -> None:
    words = text.split()
    lines: list[str] = []
    line = ""
    for word in words:
        candidate = f"{line} {word}".strip()
        if line and stringWidth(candidate, font_name, font_size) > max_width:
            lines.append(line)
            line = word
        else:
            line = candidate
    if line:
        lines.append(line)

    pdf.setFillColor(color)
    pdf.setFont(font_name, font_size)
    for index, line_text in enumerate(lines[:3]):
        pdf.drawCentredString(x_center, first_y - index * line_gap, line_text)


def draw_logo(pdf: canvas.Canvas, logo_path: Path) -> None:
    with Image.open(logo_path) as source:
        logo = crop_transparent_border(source)

    pdf.setFillColor(BG_FILL)
    pdf.rect(72, 400, 172, 132, stroke=0, fill=1)

    max_w, max_h = 132, 116
    ratio = min(max_w / logo.width, max_h / logo.height)
    draw_w = logo.width * ratio
    draw_h = logo.height * ratio
    x = 98 + (max_w - draw_w) / 2
    y = 409 + (max_h - draw_h) / 2
    pdf.drawImage(ImageReader(logo), x, y, width=draw_w, height=draw_h, mask="auto")


def draw_main_text(
    pdf: canvas.Canvas,
    attendee_name: str,
    event_title: str,
    sessions_total: str,
    uppercase_names: bool,
) -> None:
    pdf.setFillColor(BG_FILL)
    pdf.rect(145, 192, 552, 152, stroke=0, fill=1)

    rendered_name = attendee_name.upper() if uppercase_names else attendee_name
    draw_centered_fit(
        pdf,
        rendered_name,
        x_center=421,
        y=292,
        max_width=540,
        font_name="Helvetica",
        start_size=38,
        min_size=22,
        color=BROWN,
    )

    pdf.setFillColor(LIGHT_BROWN)
    pdf.setFont("Helvetica", 16)
    pdf.drawCentredString(
        421,
        248,
        f"for successfully attending all {sessions_total} sessions of the event titled",
    )

    draw_centered_fit(
        pdf,
        event_title,
        x_center=421,
        y=213,
        max_width=490,
        font_name="Helvetica-Bold",
        start_size=22,
        min_size=16,
        color=LIGHT_BROWN,
    )


def draw_signature_block(pdf: canvas.Canvas, signer_name: str, signer_title: str) -> None:
    pdf.setFillColor(BG_FILL)
    pdf.rect(555, 66, 178, 130, stroke=0, fill=1)

    signature = make_signature_image(signer_name)
    pdf.drawImage(ImageReader(signature), 565, 142, width=160, height=42, mask="auto")

    pdf.setStrokeColor(colors.HexColor("#5f5f5f"))
    pdf.setLineWidth(1.1)
    pdf.line(565, 135, 725, 135)

    pdf.setFillColor(BROWN)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(645, 116, signer_name)

    draw_wrapped_centered(
        pdf,
        signer_title,
        x_center=645,
        first_y=100,
        max_width=135,
        font_name="Helvetica-Bold",
        font_size=7.6,
        color=BROWN,
        line_gap=9,
    )


def draw_verification_block(pdf: canvas.Canvas, signed_certificate: SignedCertificate) -> None:
    qr_image = make_qr_image(signed_certificate.verify_url)
    pdf.setFillColor(BG_FILL)
    pdf.rect(727, 70, 65, 118, stroke=0, fill=1)
    pdf.drawImage(ImageReader(qr_image), 735, 119, width=48, height=48, mask="auto")

    pdf.setFillColor(INK)
    pdf.setFont("Helvetica-Bold", 5.2)
    pdf.drawCentredString(759, 108, "VERIFY")
    pdf.setFont("Helvetica", 4.5)
    pdf.drawCentredString(759, 100, "CERTIFICATE ID")
    pdf.setFont("Helvetica-Bold", 4.8)
    cert_id = signed_certificate.cert_id
    pdf.drawCentredString(759, 91, cert_id[:18])
    pdf.drawCentredString(759, 84, cert_id[18:])


def make_overlay_pdf(
    width: float,
    height: float,
    logo_path: Path,
    signed_certificate: SignedCertificate,
    sessions_total: str,
    signer_name: str,
    signer_title: str,
    uppercase_names: bool,
) -> io.BytesIO:
    packet = io.BytesIO()
    pdf = canvas.Canvas(packet, pagesize=(width, height))
    draw_logo(pdf, logo_path)
    draw_main_text(
        pdf,
        signed_certificate.attendee.name,
        signed_certificate.payload["event"],
        sessions_total,
        uppercase_names,
    )
    draw_signature_block(pdf, signer_name, signer_title)
    draw_verification_block(pdf, signed_certificate)
    pdf.save()
    packet.seek(0)
    return packet


def render_certificate(
    template_path: Path,
    logo_path: Path,
    signed_certificate: SignedCertificate,
    sessions_total: str,
    signer_name: str,
    signer_title: str,
    uppercase_names: bool,
) -> bytes:
    reader = PdfReader(str(template_path))
    page = reader.pages[0]
    width = float(page.mediabox.width)
    height = float(page.mediabox.height)

    overlay_pdf = make_overlay_pdf(
        width=width,
        height=height,
        logo_path=logo_path,
        signed_certificate=signed_certificate,
        sessions_total=sessions_total,
        signer_name=signer_name,
        signer_title=signer_title,
        uppercase_names=uppercase_names,
    )
    overlay_reader = PdfReader(overlay_pdf)
    page.merge_page(overlay_reader.pages[0])

    output = io.BytesIO()
    writer = PdfWriter()
    writer.add_page(page)
    writer.write(output)
    return output.getvalue()


def write_png(pdf_bytes: bytes, output_path: Path) -> None:
    images = convert_from_bytes(pdf_bytes, dpi=220, first_page=1, last_page=1)
    if not images:
        raise RuntimeError("Unable to render PDF to PNG.")
    images[0].save(output_path, "PNG")


def write_manifests(records: list[dict[str, str]], output_dir: Path) -> tuple[Path, Path]:
    json_path = output_dir / "manifest.json"
    csv_path = output_dir / "manifest.csv"

    json_path.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    fieldnames = [
        "cert_id",
        "name",
        "rollno",
        "email",
        "email_sha256",
        "event",
        "sessions_total",
        "issuer_name",
        "issuer_title",
        "issue_date",
        "certificate_type",
        "key_id",
        "token",
        "verify_url",
        "pdf_path",
        "png_path",
    ]
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    return csv_path, json_path


def write_public_key_json(public_key_path: Path, key_id: str, public_pem: str) -> Path:
    public_json_path = public_key_path.with_suffix(".json")
    public_json_path.write_text(
        json.dumps({"key_id": key_id, "algorithm": "Ed25519", "public_key_pem": public_pem}, indent=2),
        encoding="utf-8",
    )
    return public_json_path


def write_certificates(
    attendees: Iterable[Attendee],
    template_path: Path,
    logo_path: Path,
    output_dir: Path,
    verify_base_url: str,
    default_event_title: str,
    sessions_total: str,
    signer_name: str,
    signer_title: str,
    certificate_type: str,
    issue_date: str,
    private_key_path: Path,
    public_key_path: Path,
    combined_name: str,
    uppercase_names: bool,
) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir = output_dir / "individual" / "pdf"
    png_dir = output_dir / "individual" / "png"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    png_dir.mkdir(parents=True, exist_ok=True)

    private_key = load_or_create_private_key(private_key_path, public_key_path)
    public_key = private_key.public_key()
    key_id = key_id_for_public_key(public_key)
    public_pem = public_key_pem(public_key)
    public_json_path = write_public_key_json(public_key_path, key_id, public_pem)

    combined_writer = PdfWriter()
    pdf_files: list[Path] = []
    png_files: list[Path] = []
    records: list[dict[str, str]] = []

    for attendee in attendees:
        signed_certificate = sign_certificate(
            attendee=attendee,
            private_key=private_key,
            key_id=key_id,
            verify_base_url=verify_base_url,
            default_event_title=default_event_title,
            sessions_total=sessions_total,
            signer_name=signer_name,
            signer_title=signer_title,
            certificate_type=certificate_type,
            issue_date=issue_date,
        )
        pdf_bytes = render_certificate(
            template_path=template_path,
            logo_path=logo_path,
            signed_certificate=signed_certificate,
            sessions_total=sessions_total,
            signer_name=signer_name,
            signer_title=signer_title,
            uppercase_names=uppercase_names,
        )

        base_name = certificate_filename(attendee)
        pdf_path = pdf_dir / f"{base_name}.pdf"
        png_path = png_dir / f"{base_name}.png"
        pdf_path.write_bytes(pdf_bytes)
        write_png(pdf_bytes, png_path)
        pdf_files.append(pdf_path)
        png_files.append(png_path)

        combined_reader = PdfReader(io.BytesIO(pdf_bytes))
        combined_writer.add_page(combined_reader.pages[0])

        records.append(
            {
                "cert_id": signed_certificate.cert_id,
                "name": attendee.name,
                "rollno": attendee.rollno,
                "email": attendee.email,
                "email_sha256": signed_certificate.payload["email_sha256"],
                "event": signed_certificate.payload["event"],
                "sessions_total": signed_certificate.payload["sessions_total"],
                "issuer_name": signed_certificate.payload["issuer_name"],
                "issuer_title": signed_certificate.payload["issuer_title"],
                "issue_date": signed_certificate.payload["issue_date"],
                "certificate_type": signed_certificate.payload["certificate_type"],
                "key_id": signed_certificate.key_id,
                "token": signed_certificate.token,
                "verify_url": signed_certificate.verify_url,
                "pdf_path": str(pdf_path),
                "png_path": str(png_path),
            }
        )

    combined_path = output_dir / combined_name
    with combined_path.open("wb") as handle:
        combined_writer.write(handle)

    manifest_csv_path, manifest_json_path = write_manifests(records, output_dir)

    return {
        "pdf_files": pdf_files,
        "png_files": png_files,
        "combined_path": combined_path,
        "manifest_csv_path": manifest_csv_path,
        "manifest_json_path": manifest_json_path,
        "public_key_path": public_key_path,
        "public_json_path": public_json_path,
        "key_id": key_id,
    }


def main() -> None:
    args = parse_args()

    attendees = read_attendees(args.input_path)
    result = write_certificates(
        attendees=attendees,
        template_path=args.template,
        logo_path=args.logo,
        output_dir=args.output,
        verify_base_url=normalize_base_url(args.verify_base_url),
        default_event_title=args.event,
        sessions_total=str(args.sessions_total),
        signer_name=args.signer_name,
        signer_title=args.signer_title,
        certificate_type=args.certificate_type,
        issue_date=args.issue_date,
        private_key_path=args.private_key,
        public_key_path=args.public_key,
        combined_name=args.combined_name,
        uppercase_names=args.uppercase_names,
    )

    print(f"Generated {len(result['pdf_files'])} PDF certificate(s).")
    print(f"Generated {len(result['png_files'])} PNG certificate(s).")
    print(f"PDFs: {result['pdf_files'][0].parent}")
    print(f"PNGs: {result['png_files'][0].parent}")
    print(f"Combined PDF: {result['combined_path']}")
    print(f"Manifest CSV: {result['manifest_csv_path']}")
    print(f"Manifest JSON: {result['manifest_json_path']}")
    print(f"Public key: {result['public_key_path']}")
    print(f"Public key JSON: {result['public_json_path']}")
    print(f"Key ID: {result['key_id']}")


if __name__ == "__main__":
    main()
